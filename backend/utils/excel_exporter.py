"""
Exportador Excel no padrão visual da Internal Data Platform.
"""
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


VERDE_PRIMARIO = "004632"
ZEBRA_CINZA = "F2F2F2"
BORDA_CINZA = "CCCCCC"


def _borda():
    lado = Side(style="thin", color=BORDA_CINZA)
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def exportar_xlsx(dados: list[dict], colunas: list[tuple[str, str]] | None = None) -> bytes:
    """
    Gera arquivo .xlsx no padrão Internal Data Platform.

    dados: lista de dicts
    colunas: lista de (chave, titulo_coluna) — se None usa as chaves do dict
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"

    if not dados:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    if colunas is None:
        colunas = [(k, k.replace("_", " ").title()) for k in dados[0].keys()]

    cab_fill = PatternFill("solid", fgColor=VERDE_PRIMARIO)
    cab_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    dado_font = Font(name="Arial", size=9)
    zebra_fill = PatternFill("solid", fgColor=ZEBRA_CINZA)
    borda = _borda()
    centro = Alignment(horizontal="center", vertical="center")

    # Cabeçalho
    for col_idx, (_, titulo) in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_idx, value=titulo.upper())
        cell.font = cab_font
        cell.fill = cab_fill
        cell.border = borda
        cell.alignment = centro

    # Dados
    for row_idx, row in enumerate(dados, 2):
        fill = zebra_fill if row_idx % 2 == 0 else None
        for col_idx, (chave, _) in enumerate(colunas, 1):
            valor = row.get(chave, "")
            if valor is None:
                valor = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=str(valor))
            cell.font = dado_font
            cell.border = borda
            if fill:
                cell.fill = fill

    # Congelar cabeçalho e auto-filtro
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Ajustar largura
    for col in ws.columns:
        tamanho = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(tamanho + 4, 45)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exportar_csv(cpfs: list[str], separador: str = "\n") -> bytes:
    """Gera arquivo CSV com um CPF por linha."""
    conteudo = separador.join(cpfs)
    return conteudo.encode("utf-8-sig")


def exportar_txt(cpfs: list[str]) -> bytes:
    """Gera arquivo TXT com um CPF por linha."""
    conteudo = "\n".join(cpfs)
    return conteudo.encode("utf-8")


def exportar_xlsx_cpfs(cpfs: list[str], com_pontuacao: bool = False) -> bytes:
    """Gera XLSX somente com coluna de CPFs."""
    from utils.cpf_utils import formatar_cpf
    dados = []
    for cpf in cpfs:
        dados.append({"cpf": formatar_cpf(cpf) if com_pontuacao else cpf})
    return exportar_xlsx(dados, colunas=[("cpf", "CPF")])

